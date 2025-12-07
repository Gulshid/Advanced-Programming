# --------------------------------------------
# TEXT FILE HANDLING
# --------------------------------------------

# Write file
file = open("students.txt", "w")
file.write("Name: Ali\n")
file.write("Age: 21\n")
file.write("Department: CS\n")
file.close()

# Read entire file
file = open("students.txt", "r")
print(file.read())
file.close()

# Read line by line
file = open("students.txt", "r")
for line in file:
    print(line.strip())
file.close()

# Remove specific line
file = open("data.txt", "r")
lines = file.readlines()
file.close()
line_num = int(input("Line to delete: "))
if 1 <= line_num <= len(lines):
    del lines[line_num - 1]
file = open("data.txt", "w")
file.writelines(lines)
file.close()

# Insert new line
file = open("data.txt", "r")
lines = file.readlines()
file.close()
pos = int(input("Insert at line: "))
new = input("New line: ") + "\n"
if 1 <= pos <= len(lines) + 1:
    lines.insert(pos - 1, new)
file = open("data.txt", "w")
file.writelines(lines)
file.close()

# Append + read + count
with open("newfile.txt", "a") as f:
    f.write("Extra line\n")
    f.write("Python is powerful\n")

with open("newfile.txt", "r") as f:
    content = f.readlines()

print(content)
print("Lines:", len(content))
words = sum(len(line.split()) for line in content)
print("Words:", words)

# --------------------------------------------
# EXCEL (openpyxl)
# --------------------------------------------

import openpyxl

# Create Excel
wb = openpyxl.Workbook()
sh = wb.active
sh["A1"] = "Name"; sh["B1"] = "Marks"; sh["C1"] = "Dept"
sh["A2"] = "Ali"; sh["B2"] = 85; sh["C2"] = "CS"
sh["A3"] = "Ayesha"; sh["B3"] = 90; sh["C3"] = "IT"
wb.save("students.xlsx")

# Read Excel
wb = openpyxl.load_workbook("students.xlsx")
sh = wb.active
print(sh["A2"].value, sh["B2"].value, sh["C2"].value)

# Read all rows
for row in sh.iter_rows(min_row=2, values_only=True):
    print(row)

# Append new row
wb = openpyxl.load_workbook("students.xlsx")
sh = wb.active
r = sh.max_row + 1
sh[f"A{r}"] = "Hamza"
sh[f"B{r}"] = 80
sh[f"C{r}"] = "AI"
wb.save("students.xlsx")

# Update cell
wb = openpyxl.load_workbook("students.xlsx")
sh = wb.active
sh["B2"] = 95
wb.save("students.xlsx")

# --------------------------------------------
# EXCEL USING PANDAS
# --------------------------------------------

import pandas as pd

# Write Excel
data = {
    "Name": ["Ali", "Ayesha", "Sara", "Usman"],
    "Marks": [85, 90, 88, 92],
    "Department": ["CS", "IT", "SE", "CS"]
}
df = pd.DataFrame(data)
df.to_excel("results.xlsx", index=False)

# Read Excel
df = pd.read_excel("results.xlsx")
print(df)
